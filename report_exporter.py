# report_exporter.py

import calendar
from pathlib import Path
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from svatky import ceske_statni_svatky
from calculator import MzdovySystem

# České názvy měsíců (1–12)
MESICE_CZ = [
    None, "Leden", "Únor", "Březen", "Duben", "Květen", "Červen",
    "Červenec", "Srpen", "Září", "Říjen", "Listopad", "Prosinec"
]

# šedá výplň pro nepracovní dny
GREY_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

def set_cell_value(ws, coord: str, value):
    """
    Zapíše `value` do buňky `coord`.
    Pokud je `coord` ve sloučeném rozsahu, vloží do levé‑horní buňky toho rozsahu.
    """
    col_letter, row_str = coordinate_from_string(coord)
    row = int(row_str)
    col = column_index_from_string(col_letter)
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            ws.cell(row=merged.min_row, column=merged.min_col).value = value
            return
    ws[coord].value = value

def set_row_fill(ws, row: int, cols: list[str], fill: PatternFill):
    """Aplikuje `fill` na všechny buňky zvolených sloupců v řádku `row`."""
    for col in cols:
        ws[f"{col}{row}"].fill = fill

def export_vykaz(
    zamestnanec,
    mesic: int,
    rok: int,
    template_path: Path,
    out_path: Path,
    start_time: str
) -> Path:
    """
    Vygeneruje Excel podle šablony:
      - do E2: PERIODA (Měsíc Rok)
      - do C3: Jméno a příjmení
      - do C4: Typ smlouvy
      - od řádku 7 (sloupce A–F):
          A = den v měsíci (u svátků "číslo SV")
          B = začátek pracovní doby (jen pro pracovní dny s>0)
          C = konec pracovní doby (jen pro pracovní dny s>0)
          D, E = "-" pro přestávku (jen pro pracovní dny s>0)
          F = počet odpracovaných hodin
      - celý řádek pro svátky a víkendy zašedí
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # připravíme systém pro kontrolu pracovních dnů
    svatky = ceske_statni_svatky(rok)
    system = MzdovySystem(svatky=svatky)

    # PERIODA, JMÉNO, TYP SMLOUVY
    set_cell_value(ws, "E2", f"{MESICE_CZ[mesic]} {rok}")
    set_cell_value(ws, "C3", f"{zamestnanec.jmeno} {zamestnanec.prijmeni}")
    set_cell_value(ws, "C4", zamestnanec.typ_smlouvy)

    # Parsování start_time
    h, m = map(int, start_time.split(":"))
    base_time = timedelta(hours=h, minutes=m)

    # Pro každý den v měsíci
    start_row = 7
    max_day = calendar.monthrange(rok, mesic)[1]
    for day in range(1, max_day + 1):
        row = start_row + day - 1
        datum = date(rok, mesic, day)

        # 1) vypíšeme den, případně s "SV"
        if datum in svatky:
            set_cell_value(ws, f"A{row}", f"{day} SV")
        else:
            set_cell_value(ws, f"A{row}", day)

        # 2) nepracovní dny (svátky nebo víkend) zašedíme a pokračujeme
        if datum in svatky or not system.je_pracovni_den(datum):
            set_row_fill(ws, row, ["A","B","C","D","E","F"], GREY_FILL)
            continue

        # 3) pracovní den – data podle odpracovaných hodin
        hod = zamestnanec.odpracovane_hodiny.get(datum, 0.0)
        if hod > 0:
            set_cell_value(ws, f"B{row}", start_time)
            end_dt = base_time + timedelta(hours=hod)
            end_str = f"{int(end_dt.total_seconds()//3600)}:{int((end_dt.total_seconds()%3600)//60):02d}"
            set_cell_value(ws, f"C{row}", end_str)
            set_cell_value(ws, f"D{row}", "-")
            set_cell_value(ws, f"E{row}", "-")
        set_cell_value(ws, f"F{row}", hod)

    wb.save(out_path)
    return out_path
